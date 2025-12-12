# Cargon9003 - Final Exam
# 12/09/2025

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
import matplotlib.pyplot as plt
from datetime import datetime

# Auto-create and go to C:\FinalExam
os.makedirs(r"C:\FinalExam", exist_ok=True)
os.chdir(r"C:\FinalExam")

print("Cargon9003")

def askUser():
    total = 0
    print("\nPlease enter 5 numbers:")
    for i in range(1, 6):
        num = int(input(f"Number {i}: "))
        total += num
    print(f"The sum for the 5 numbers entered is: {total}")

def askIncome():
    print("\nEnter 5 people:")
    with open("final.csv", "a", encoding="utf-8") as f:
        for i in range(5):
            name = input("Name: ")
            income = input("Income: ")
            f.write(f"\n{name},{income}")

def excelPie():
    df = pd.read_csv("final.csv", header=None, names=["Name","Income"])
    df["Income"] = pd.to_numeric(df["Income"], errors='coerce').astype(int)

    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Income"])
    for _, row in df.iterrows():
        ws.append([row["Name"], row["Income"]])

    pie = PieChart()
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    pie.title = f"Cargon9003 {datetime.now().strftime('%B %d, %Y')}"

    ws.add_chart(pie, "D2")
    wb.save("final.xlsx")          # THIS LINE WAS MISSING BEFORE
    print("Pie chart saved → final.xlsx")

def verticalBar():
    df = pd.read_csv("final.csv", header=None, names=["Name","Income"])
    df["Income"] = pd.to_numeric(df["Income"], errors='coerce').astype(int)

    plt.figure(figsize=(10,6))
    plt.bar(df["Name"], df["Income"], color='green')
    plt.title(f"Cargon9003 {datetime.now().strftime('%B %d, %Y')}")
    plt.xlabel("Name")
    plt.ylabel("Income")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.show()

# RUN EVERYTHING
askUser()
askIncome()
excelPie()        # This now creates final.xlsx
verticalBar()